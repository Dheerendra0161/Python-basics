import openpyxl
from faker import Faker

# Provide the location\Path, if you will give the name of the file only it will generate at the project location only
# file_name = 'C:\\Users\\Dheerendra\\Desktop\\Python\\ExcelFilePython1.xlsx'

file_name = 'ExcelFile'

# Try to load the workbook, create a new one if it doesn't exist
try:
    workbook = openpyxl.load_workbook(file_name + '.xlsx')
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# Create a new sheet or select an existing one.
sheet = workbook.create_sheet("New_Test_Data", 0)

# Set headers
headers = ['Name', 'Email', 'Address']
sheet.append(headers)

# Create a Faker instance
fake = Faker()

# Get the next available row
next_row = sheet.max_row + 1

# Populate the sheet with fake data
for i in range(next_row, next_row + 10):  # 10 rows of data
    # For each row, generate fake data for three columns
    for j in range(1, 4):  # Three columns
        # You can generate different fake data for each column
        if j == 1:
            sheet.cell(row=i, column=j, value=fake.name())
        elif j == 2:
            sheet.cell(row=i, column=j, value=fake.email())
        elif j == 3:
            sheet.cell(row=i, column=j, value=fake.address())

# Save the workbook
workbook.save(file_name + '.xlsx')
print(f"Excel file '{file_name}.xlsx' has been created and populated with fake data.")
