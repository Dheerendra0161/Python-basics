import openpyxl

file_name = 'C:\\Users\\Dheerendra\\Desktop\\Python\\ExcelFilePython.xlsx'
workbook = openpyxl.load_workbook(file_name)
sheet = workbook['TestData']
total_row = sheet.max_row
total_col = sheet.max_column

for r in range(1, total_row + 1):
    for c in range(1, total_col + 1):
        print(sheet.cell(r, c).value,
              end=' ')  # end=' ', it will end the print statement with a space instead of a newline.
    print()

print(f"Excel file '{file_name}' has been created.")
