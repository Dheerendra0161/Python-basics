import openpyxl

file_name = 'C:\\Users\\Dheerendra\\Desktop\\Python\\ExcelFilePython.xlsx'
workbook = openpyxl.load_workbook(file_name)
# Select the sheet named 'TestData'
sheet = workbook['TestData']
# sheet=workbook.active     #current open active sheet in excel file

# Data to write
data = [
    ('Alice', 30, 'Engineer'),  # Tuple is an ordered and immutable collection of elements.()
    ('Bob', 25, 'Manager'),
    ('Charlie', 35, 'Data Scientist')
]

# Find the next empty row for writing data
next_row = sheet.max_row + 1
# Write new data to the sheet
for data_row in data:
    for col, value in enumerate(data_row, start=1):
        # start=1 specifies that the index should start at 1 instead of the default 0.
        sheet.cell(row=next_row, column=col, value=value)
    next_row += 1

# Save the workbook
workbook.save(file_name)
print(f"Data has been written to Excel file '{file_name}'.")
